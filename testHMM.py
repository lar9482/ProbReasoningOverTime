from HMM.SleepMM.SleepMM import getSleepMM
from HMM.SleepMM.SleepMMVariables import SleepInClass_E, RedEyes_E
from HMM.countryDance import countryDance
from HMM.fixedLagSmoothing import FixedLagSmoothing
from HMM.viterbiAlgorithm import viterbiAlgorithm

import numpy as np
import random

from enum import Enum

from openpyxl import Workbook

class HMMEvidence(Enum):
    redEyes_sleepClass = 0
    redEyes_notSleepClass = 1
    notRedEyes_sleepClass = 2
    notRedEyes_notSleepClass = 3
    random = 4

class fixedLagSmoothingTestResult():
    def __init__(self):
        self.t = 0

        self.t2CountryDance_0 = ''
        self.t2FixedLag_0 = ''

        self.t3CountryDance_0 = ''
        self.t3FixedLag_0 = ''

        self.t4CountryDance_0 = ''
        self.t4FixedLag_0 = ''

        self.t5CountryDance_0 = ''
        self.t5FixedLag_0 = ''

class likelyPathTestResult():
    def __init__(self):
        self.t = 0
        self.distroFrom1ToT_0 = []
        self.mostLikelyPath = []

def testHMM(HMMEvidenceOption):
    sleepMM = getSleepMM()

    time = 25
    completeEvidenceList = resolveEvidenceOption(HMMEvidenceOption, time)

    twoLagSmoothing = FixedLagSmoothing(sleepMM, 2)
    threeLagSmoothing = FixedLagSmoothing(sleepMM, 3)
    fourLagSmoothing = FixedLagSmoothing(sleepMM, 4)
    fiveLagSmoothing = FixedLagSmoothing(sleepMM, 5)

    allSmoothingResults = []
    allLikelyPathResults = []
    allStateEstimationResults = []

    for t in range(1, time+1):
        # Padding the evidenceList with a value at time 0
        # in order to begin indexing at time 1.
        evidenceFrom1ToT = ["noEvidenceAt0"] + completeEvidenceList[1:t+1]

        distroFrom1ToT = countryDance(sleepMM, evidenceFrom1ToT, t)
        distroAt_T2 = twoLagSmoothing.runSmoothing(completeEvidenceList[t])
        distroAt_T3 = threeLagSmoothing.runSmoothing(completeEvidenceList[t])
        distroAt_T4 = fourLagSmoothing.runSmoothing(completeEvidenceList[t])
        distroAt_T5 = fiveLagSmoothing.runSmoothing(completeEvidenceList[t])
        mostLikelyPath = viterbiAlgorithm(sleepMM, evidenceFrom1ToT)
        
        smoothingResult = fixedLagSmoothingTestResult()
        smoothingResult.t = t
        if (isinstance(distroAt_T2, np.ndarray)):
            smoothingResult.t2CountryDance_0 = round(distroFrom1ToT[t-twoLagSmoothing.lagValue-1][0][0], 4)
            smoothingResult.t2FixedLag_0 = round(distroAt_T2[0][0], 4)

        if (isinstance(distroAt_T3, np.ndarray)):
            smoothingResult.t3CountryDance_0 = round(distroFrom1ToT[t-threeLagSmoothing.lagValue-1][0][0], 4)
            smoothingResult.t3FixedLag_0 = round(distroAt_T3[0][0], 4)

        if(isinstance(distroAt_T4, np.ndarray)):
            smoothingResult.t4CountryDance_0 = round(distroFrom1ToT[t-fourLagSmoothing.lagValue-1][0][0], 4)
            smoothingResult.t4FixedLag_0 = round(distroAt_T4[0][0], 4)

        if(isinstance(distroAt_T5, np.ndarray)):
            smoothingResult.t5CountryDance_0 = round(distroFrom1ToT[t-fiveLagSmoothing.lagValue-1][0][0], 4)
            smoothingResult.t5FixedLag_0 = round(distroAt_T5[0][0], 4)
        

        mostLikelyPathResult = likelyPathTestResult()
        mostLikelyPathResult.t = t
        for distro in distroFrom1ToT:
            mostLikelyPathResult.distroFrom1ToT_0.append(round(distro[0][0], 4))
        
        mostLikelyPathResult.mostLikelyPath = mostLikelyPath

        allStateEstimationResults.append((t, round(distroFrom1ToT[t-1][0][0], 4)))
        allLikelyPathResults.append(mostLikelyPathResult)
        allSmoothingResults.append(smoothingResult)

    saveSmoothingResults(allSmoothingResults, HMMEvidenceOption.name)
    saveMostLikelyPathResults(allLikelyPathResults, HMMEvidenceOption.name)
    saveStateEstimationResults(allStateEstimationResults, HMMEvidenceOption.name)

def saveSmoothingResults(allSmoothingResults, testName):
    workbook = Workbook()
    sheet = workbook.active

    # Add headers
    sheet.append(['EnoughSleep'])
    sheet.append(['T', 
                  'CountryDance_T-2', 'FixedLag_T-2', 
                  'CountryDance_T-3', 'FixedLag_T-3',
                  'CountryDance_T-4', 'FixedLag_T-4', 
                  'CountryDance_T-5', 'FixedLag_T-5'])
    for testResult in allSmoothingResults:
        sheet.append([
            testResult.t,

            testResult.t2CountryDance_0,
            testResult.t2FixedLag_0,

            testResult.t3CountryDance_0,
            testResult.t3FixedLag_0,

            testResult.t4CountryDance_0,
            testResult.t4FixedLag_0,

            testResult.t5CountryDance_0,
            testResult.t5FixedLag_0
        ])
        
    workbook.save('./results/HMM/' + testName + '_smoothingResults.xlsx')

def saveMostLikelyPathResults(allLikelyPathResults, testName):
    workbook = Workbook()
    sheet = workbook.active

    # Add headers
    sheet.append(['T', 'CountryDanceFrom1ToT_GivenEnoughSleep', 'mostLikelyPathFrom1ToT',])
    for testResult in allLikelyPathResults:
        distributionFrom1ToT = str(testResult.distroFrom1ToT_0)
        likelyPath = str(testResult.mostLikelyPath)

        sheet.append([
            testResult.t,
            distributionFrom1ToT,
            likelyPath,
        ])

    workbook.save('./results/HMM/' + testName + '_pathingResults.xlsx')

def saveStateEstimationResults(allStateEstimationResults, testName):
    workbook = Workbook()
    sheet = workbook.active

    # Add headers
    sheet.append(['EnoughSleep'])
    sheet.append(['T', 'Prob', 'state'])

    for timeAndProb in allStateEstimationResults:
        t = timeAndProb[0]
        prob = timeAndProb[1]

        stateEstimate = ''
        if (prob < 0.5):
            stateEstimate = 'notEnoughSleep'
        else:
            stateEstimate = 'enoughSleep'

        sheet.append([
            t,
            prob,
            stateEstimate
        ])

    workbook.save('./results/HMM/' + testName + '_stateEstimation.xlsx')
    
def resolveEvidenceOption(HMMEvidenceOption, time):
    allEvidenceTuple = [
        (RedEyes_E.redEyes.value, SleepInClass_E.sleepInClass.value),
        (RedEyes_E.redEyes.value, SleepInClass_E.notSleepInClass.value),
        (RedEyes_E.notRedEyes.value, SleepInClass_E.sleepInClass.value),
        (RedEyes_E.notRedEyes.value, SleepInClass_E.notSleepInClass.value)
    ]

    if (HMMEvidenceOption.value == HMMEvidence.redEyes_sleepClass.value):
        return [(RedEyes_E.redEyes.value, SleepInClass_E.sleepInClass.value) for _ in range(0, time+1)]

    elif (HMMEvidenceOption.value == HMMEvidence.redEyes_notSleepClass.value):
        return [(RedEyes_E.redEyes.value, SleepInClass_E.notSleepInClass.value) for _ in range(0, time+1)]
    
    elif (HMMEvidenceOption.value == HMMEvidence.notRedEyes_sleepClass.value):
        return [(RedEyes_E.notRedEyes.value, SleepInClass_E.sleepInClass.value) for _ in range(0, time+1)]

    elif (HMMEvidenceOption.value == HMMEvidence.notRedEyes_notSleepClass.value):
        return [(RedEyes_E.notRedEyes.value, SleepInClass_E.notSleepInClass.value) for _ in range(0, time+1)]
    
    elif (HMMEvidenceOption.value == HMMEvidence.random.value):
        return [random.choice(allEvidenceTuple) for _ in range(0, time+1)]
    else:
        return [random.choice(allEvidenceTuple) for _ in range(0, time+1)]